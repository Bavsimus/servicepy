import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QComboBox, QMessageBox, QFileDialog, QSpinBox, QFormLayout, QGroupBox,
    QHBoxLayout, QInputDialog
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from calendar import monthrange
from datetime import datetime
import random

GUZERGAH_DOSYASI = "guzergahlar.txt"

def guzergah_listesini_yukle():
    if not os.path.exists(GUZERGAH_DOSYASI):
        return []
    with open(GUZERGAH_DOSYASI, 'r', encoding='utf-8') as f:
        return [satir.strip() for satir in f.readlines() if satir.strip()]

def guzergah_ekle(guzergah):
    guzergahlar = guzergah_listesini_yukle()
    if guzergah not in guzergahlar:
        guzergahlar.append(guzergah)
        with open(GUZERGAH_DOSYASI, 'w', encoding='utf-8') as f:
            f.write("\n".join(guzergahlar))

class ExcelGenerator(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Excel Tablo Oluşturucu (Güzergah Destekli)')
        self.setGeometry(100, 100, 800, 600)
        self.layout = QVBoxLayout()

        self.combo_yil = QComboBox()
        self.combo_yil.addItems([str(yil) for yil in range(2020, 2031)])

        self.combo_ay = QComboBox()
        self.combo_ay.addItems([
            'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
            'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'
        ])

        self.spin_arac_sayisi = QSpinBox()
        self.spin_arac_sayisi.setRange(1, 10)
        self.spin_arac_sayisi.valueChanged.connect(self.arac_formlarini_guncelle)

        guzergah_btn_layout = QHBoxLayout()
        guzergah_ekle_btn = QPushButton("+ Güzergah Ekle")
        guzergah_ekle_btn.clicked.connect(self.guzergah_ekle)
        guzergah_btn_layout.addWidget(guzergah_ekle_btn)

        self.layout.addWidget(QLabel("Yıl Seçin:"))
        self.layout.addWidget(self.combo_yil)
        self.layout.addWidget(QLabel("Ay Seçin:"))
        self.layout.addWidget(self.combo_ay)
        self.layout.addWidget(QLabel("Araç Sayısı:"))
        self.layout.addWidget(self.spin_arac_sayisi)
        self.layout.addLayout(guzergah_btn_layout)

        self.arac_formlari_container = QVBoxLayout()
        self.arac_formlari_group = QGroupBox("Araç Bilgileri")
        self.arac_formlari_group.setLayout(self.arac_formlari_container)
        self.layout.addWidget(self.arac_formlari_group)

        self.button_kaydet = QPushButton("Excel'i Oluştur")
        self.button_kaydet.clicked.connect(self.kaydet_ve_listele)
        self.layout.addWidget(self.button_kaydet)

        self.setLayout(self.layout)
        self.arac_inputlar = []
        self.arac_formlarini_guncelle()

    def guzergah_ekle(self):
        text, ok = QInputDialog.getText(self, 'Güzergah Ekle', 'Yeni güzergah girin:')
        if ok and text.strip():
            guzergah_ekle(text.strip())
            self.arac_formlarini_guncelle()

    def arac_formlarini_guncelle(self):
        guzergahlar = guzergah_listesini_yukle()
        for i in reversed(range(self.arac_formlari_container.count())):
            self.arac_formlari_container.itemAt(i).widget().setParent(None)
        self.arac_inputlar.clear()

        for i in range(self.spin_arac_sayisi.value()):
            grup = QGroupBox(f"Araç {i + 1}")
            form = QFormLayout()

            input_baslangic_km = QLineEdit()
            input_km_aralik = QLineEdit()
            input_gorev_yeri = QComboBox()
            input_gorev_yeri.addItems(guzergahlar)
            input_haftasonu = QComboBox()
            input_haftasonu.addItems(['Çalışıyor', 'Çalışmıyor'])

            form.addRow("Gün Başı Kilometre:", input_baslangic_km)
            form.addRow("KM Aralığı (örn: 90-100):", input_km_aralik)
            form.addRow("Görev Yeri:", input_gorev_yeri)
            form.addRow("Hafta Sonu Durumu:", input_haftasonu)

            grup.setLayout(form)
            self.arac_formlari_container.addWidget(grup)

            self.arac_inputlar.append({
                "baslangic_km": input_baslangic_km,
                "km_aralik": input_km_aralik,
                "gorev_yeri": input_gorev_yeri,
                "haftasonu": input_haftasonu
            })

    def kaydet_ve_listele(self):
        ay_index = self.combo_ay.currentIndex()
        yil = int(self.combo_yil.currentText())
        gun_sayisi = monthrange(yil, ay_index + 1)[1]

        wb = Workbook()
        ws = wb.active
        ws.title = "Tüm Araçlar"
        row_cursor = 1

        fill = PatternFill(start_color='FFC0C0', end_color='FFC0C0', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for index, arac in enumerate(self.arac_inputlar):
            try:
                km_baslangic = float(arac["baslangic_km"].text())
                km_min, km_max = map(int, arac["km_aralik"].text().split('-'))
                gorev_yeri = arac["gorev_yeri"].currentText()
                haftasonu_durumu = arac["haftasonu"].currentText()
            except Exception:
                QMessageBox.warning(self, "Hata", f"Araç {index+1} için girişler eksik veya hatalı.")
                return

            # Başlık
            ws.cell(row=row_cursor, column=1).value = "TARİH"
            ws.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=3)
            ws.cell(row=row_cursor, column=2).value = "GÜN BAŞI (km)"
            ws.cell(row=row_cursor, column=4).value = "GÜN SONU (km)"
            ws.cell(row=row_cursor, column=5).value = "YAPTIĞI KİLOMETRE"
            ws.cell(row=row_cursor, column=6).value = "KONTROL / İMZA"
            ws.merge_cells(start_row=row_cursor, start_column=7, end_row=row_cursor, end_column=9)
            ws.cell(row=row_cursor, column=7).value = "GÖREVE GİDİLEN YER"

            for col in range(1, 10):
                cell = ws.cell(row=row_cursor, column=col)
                if cell.value not in (None, ""):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            row_cursor += 1
            current_km = km_baslangic

            for day in range(1, gun_sayisi + 1):
                tarih = datetime(yil, ay_index + 1, day)
                tarih_str = tarih.strftime('%d.%m.%Y')

                ws.cell(row=row_cursor, column=1).value = tarih_str
                ws.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=3)
                ws.merge_cells(start_row=row_cursor, start_column=7, end_row=row_cursor, end_column=9)

                if tarih.weekday() >= 5 and haftasonu_durumu == 'Çalışmıyor':
                    for col in range(1, 10):
                        ws.cell(row=row_cursor, column=col).fill = fill
                        ws.cell(row=row_cursor, column=col).border = border
                    row_cursor += 1
                    continue

                yapilan_km = random.randint(km_min, km_max)
                gun_sonu_km = current_km + yapilan_km

                ws.cell(row=row_cursor, column=2).value = current_km
                ws.cell(row=row_cursor, column=4).value = gun_sonu_km
                ws.cell(row=row_cursor, column=5).value = yapilan_km
                ws.cell(row=row_cursor, column=6).value = ""
                ws.cell(row=row_cursor, column=7).value = gorev_yeri

                for col in range(1, 10):
                    cell = ws.cell(row=row_cursor, column=col)
                    if cell.value not in (None, ""):
                        cell.border = border

                if tarih.weekday() >= 5:
                    for col in range(1, 10):
                        ws.cell(row=row_cursor, column=col).fill = fill

                current_km = gun_sonu_km
                row_cursor += 1

            row_cursor += 2

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", f"Arac_Raporu_{yil}_{ay_index + 1}.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb.save(file_path)
        QMessageBox.information(self, "Başarılı", "Excel dosyası başarıyla kaydedildi!")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ExcelGenerator()
    window.show()
    sys.exit(app.exec_())
